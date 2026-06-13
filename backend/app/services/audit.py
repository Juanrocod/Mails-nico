import io
from typing import Optional
from uuid import UUID

import openpyxl
from sqlalchemy.orm import Session

from app.models.audit import AuditEvent, AccionAudit


def record_event(
    orden_id: UUID,
    usuario_id: Optional[UUID],
    accion: AccionAudit,
    ip_origen: Optional[str],
    db: Session,
    detalle: Optional[dict] = None,
) -> AuditEvent:
    event = AuditEvent(
        orden_id=orden_id,
        usuario_id=usuario_id,
        accion=accion,
        ip_origen=ip_origen,
        detalle=detalle,
    )
    db.add(event)
    return event


def get_events_for_orden(orden_id: UUID, db: Session) -> list[AuditEvent]:
    return (
        db.query(AuditEvent)
        .filter(AuditEvent.orden_id == orden_id)
        .order_by(AuditEvent.timestamp.asc())
        .all()
    )


def export_audit_trail_excel(orden_id: UUID, db: Session) -> bytes:
    events = get_events_for_orden(orden_id, db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Trail"

    headers = ["Timestamp", "Acción", "Usuario ID", "IP", "Detalle"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, event in enumerate(events, 2):
        ws.cell(row=row_idx, column=1, value=str(event.timestamp) if event.timestamp else "")
        ws.cell(row=row_idx, column=2, value=event.accion.value if event.accion else "")
        ws.cell(row=row_idx, column=3, value=str(event.usuario_id) if event.usuario_id else "sistema")
        ws.cell(row=row_idx, column=4, value=event.ip_origen or "")
        ws.cell(row=row_idx, column=5, value=str(event.detalle) if event.detalle else "")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
