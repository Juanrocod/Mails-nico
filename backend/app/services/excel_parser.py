import io
from dataclasses import dataclass
from decimal import Decimal
from typing import Optional

import openpyxl


class ExcelParseError(Exception):
    pass


@dataclass
class DeudorRow:
    clave_union: str
    nombre: str
    localidad: Optional[str]
    monto: Decimal


@dataclass
class MaestroRow:
    clave_union: str
    nombre: str
    email: Optional[str]
    localidad: Optional[str]


# Aliases aceptados por columna. Agregar alias cuando el cliente comparta el Excel real.
DEUDOR_ALIASES: dict[str, list[str]] = {
    "clave_union": ["nro cliente", "nro_cliente", "cliente", "id cliente", "codigo"],
    "nombre": ["nombre", "detalle del nombre", "detalle", "razon social", "consorcio"],
    "localidad": ["localidad", "provincia", "ciudad", "zona"],
    "monto": ["monto", "deuda", "importe", "saldo", "monto adeudado"],
}

MAESTRO_ALIASES: dict[str, list[str]] = {
    "clave_union": ["nro cliente", "nro_cliente", "cliente", "id cliente", "codigo"],
    "nombre": ["nombre", "detalle del nombre", "detalle", "razon social", "consorcio"],
    "email": ["email", "mail", "correo", "e-mail"],
    "localidad": ["localidad", "provincia", "ciudad", "zona"],
}


def _normalize(s: str) -> str:
    return s.strip().lower()


def _find_column(headers: list[str], field: str, aliases: dict[str, list[str]]) -> Optional[int]:
    normalized = [_normalize(h) for h in headers]
    for alias in aliases[field]:
        if alias in normalized:
            return normalized.index(alias)
    return None


def _load_workbook(file_bytes: bytes) -> list[list]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return rows


def parse_deudores(file_bytes: bytes) -> list[DeudorRow]:
    rows = _load_workbook(file_bytes)
    if not rows:
        raise ExcelParseError("El archivo está vacío")
    headers = [str(h) if h is not None else "" for h in rows[0]]
    col = {}
    for field in DEUDOR_ALIASES:
        idx = _find_column(headers, field, DEUDOR_ALIASES)
        if idx is None and field != "localidad":
            raise ExcelParseError(f"Columna requerida no encontrada: {field}. Encabezados: {headers}")
        col[field] = idx

    result = []
    for row in rows[1:]:
        clave = str(row[col["clave_union"]] or "").strip()
        nombre = str(row[col["nombre"]] or "").strip()
        monto_raw = row[col["monto"]]
        if not clave or not nombre or monto_raw is None:
            continue
        try:
            monto = Decimal(str(monto_raw))
        except Exception:
            continue
        if monto <= 0:
            continue
        localidad = None
        if col["localidad"] is not None:
            localidad = str(row[col["localidad"]] or "").strip() or None
        result.append(DeudorRow(clave_union=clave, nombre=nombre, localidad=localidad, monto=monto))
    return result


def parse_maestro(file_bytes: bytes) -> list[MaestroRow]:
    rows = _load_workbook(file_bytes)
    if not rows:
        raise ExcelParseError("El archivo está vacío")
    headers = [str(h) if h is not None else "" for h in rows[0]]
    col = {}
    for field in MAESTRO_ALIASES:
        idx = _find_column(headers, field, MAESTRO_ALIASES)
        if idx is None and field not in ("localidad", "email"):
            raise ExcelParseError(f"Columna requerida no encontrada: {field}. Encabezados: {headers}")
        col[field] = idx

    result = []
    for row in rows[1:]:
        clave = str(row[col["clave_union"]] or "").strip()
        nombre = str(row[col["nombre"]] or "").strip()
        if not clave or not nombre:
            continue
        email = None
        if col.get("email") is not None:
            raw_email = str(row[col["email"]] or "").strip()
            email = raw_email if raw_email else None
        localidad = None
        if col.get("localidad") is not None:
            raw_loc = str(row[col["localidad"]] or "").strip()
            localidad = raw_loc if raw_loc else None
        result.append(MaestroRow(clave_union=clave, nombre=nombre, email=email, localidad=localidad))
    return result
