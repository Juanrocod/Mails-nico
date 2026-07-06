from datetime import datetime, timezone
from decimal import Decimal

from app.models.ciclo import Ciclo
from app.models.envio import Envio, EstadoEnvio
from app.models.cliente_maestro import ClienteMaestro


def _make_ciclo(db, numero):
    c = Ciclo(numero=numero, activo=False, creado_en=datetime.now(timezone.utc))
    db.add(c)
    db.flush()
    return c


def _make_envio(db, ciclo, clave, estado=EstadoEnvio.NO_CONTESTADO, monto="1000", saldado_en=None, ciclo_numero=1):
    e = Envio(
        ciclo_id=ciclo.id, ciclo_numero=ciclo_numero, clave_union=clave, nombre_consorcio=f"Cons {clave}",
        email=f"{clave}@mail.com", monto=Decimal(monto), estado=estado,
        saldado_en=saldado_en, actualizado_en=datetime.now(timezone.utc),
    )
    db.add(e)
    db.flush()
    return e


def test_marcar_saldados_marca_ausentes_y_no_presentes(db):
    from app.services.ciclo_service import marcar_saldados

    ciclo = _make_ciclo(db, 9001)
    ausente = _make_envio(db, ciclo, "SAL-A")
    presente = _make_envio(db, ciclo, "SAL-B")
    db.commit()

    count = marcar_saldados(db, ciclo.id, {"SAL-B"})
    db.commit()

    assert count == 1
    db.refresh(ausente)
    db.refresh(presente)
    assert ausente.saldado_en is not None
    assert presente.saldado_en is None

    db.query(Envio).filter(Envio.ciclo_id == ciclo.id).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id == ciclo.id).delete(synchronize_session=False)
    db.commit()


def test_marcar_saldados_cubre_todos_los_estados(db):
    from app.services.ciclo_service import marcar_saldados

    ciclo = _make_ciclo(db, 9002)
    estados = [EstadoEnvio.NO_CONTESTADO, EstadoEnvio.CONTESTADO, EstadoEnvio.PAGO,
               EstadoEnvio.REBOTADO, EstadoEnvio.SIN_EMAIL, EstadoEnvio.FILTRADO]
    envios = [_make_envio(db, ciclo, f"SAL-EST-{i}", estado=est) for i, est in enumerate(estados)]
    db.commit()

    count = marcar_saldados(db, ciclo.id, set())
    db.commit()

    assert count == len(estados)
    for e in envios:
        db.refresh(e)
        assert e.saldado_en is not None

    db.query(Envio).filter(Envio.ciclo_id == ciclo.id).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id == ciclo.id).delete(synchronize_session=False)
    db.commit()


def test_marcar_saldados_no_pisa_saldados_previos(db):
    """Un envio ya saldado en una transicion anterior conserva su fecha original."""
    from app.services.ciclo_service import marcar_saldados

    ciclo = _make_ciclo(db, 9003)
    fecha_original = datetime(2026, 1, 1, tzinfo=timezone.utc)
    ya_saldado = _make_envio(db, ciclo, "SAL-C", saldado_en=fecha_original)
    db.commit()

    count = marcar_saldados(db, ciclo.id, set())
    db.commit()

    assert count == 0
    db.refresh(ya_saldado)
    assert ya_saldado.saldado_en.year == 2026 and ya_saldado.saldado_en.month == 1

    db.query(Envio).filter(Envio.ciclo_id == ciclo.id).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id == ciclo.id).delete(synchronize_session=False)
    db.commit()


def test_racha_se_resetea_si_ultimo_envio_esta_saldado(db):
    """Un deudor que saldo (inferido) y reaparece arranca la racha de cero."""
    from app.services.excel_joiner import _ciclos_consecutivos_deudor

    ciclo = _make_ciclo(db, 9004)
    _make_envio(db, ciclo, "SAL-D", ciclo_numero=4, saldado_en=datetime.now(timezone.utc))
    db.commit()

    assert _ciclos_consecutivos_deudor(db, "SAL-D") == 0

    db.query(Envio).filter(Envio.ciclo_id == ciclo.id).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id == ciclo.id).delete(synchronize_session=False)
    db.commit()


def test_racha_usa_el_envio_mas_reciente_no_el_de_racha_maxima(db):
    """El orden es por Ciclo.numero (mas reciente), no por Envio.ciclo_numero (racha maxima).

    Historial: en el ciclo 9601 el deudor llego a racha 5 y saldo; en el ciclo
    9602 (posterior) volvio a deber con racha 1 sin saldar. Con el ordenamiento
    viejo (Envio.ciclo_numero desc) se tomaria el envio de racha 5 — y como esta
    saldado devolveria 0. El correcto toma el envio del ciclo mas reciente: 1.
    """
    from app.services.excel_joiner import _ciclos_consecutivos_deudor

    ciclo_viejo = _make_ciclo(db, 9601)
    ciclo_reciente = _make_ciclo(db, 9602)
    _make_envio(db, ciclo_viejo, "SAL-REC", ciclo_numero=5, saldado_en=datetime.now(timezone.utc))
    _make_envio(db, ciclo_reciente, "SAL-REC", ciclo_numero=1)
    db.commit()

    assert _ciclos_consecutivos_deudor(db, "SAL-REC") == 1

    db.query(Envio).filter(Envio.ciclo_id.in_([ciclo_viejo.id, ciclo_reciente.id])).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id.in_([ciclo_viejo.id, ciclo_reciente.id])).delete(synchronize_session=False)
    db.commit()


def test_racha_sigue_si_ultimo_envio_no_saldado(db):
    from app.services.excel_joiner import _ciclos_consecutivos_deudor

    ciclo = _make_ciclo(db, 9005)
    _make_envio(db, ciclo, "SAL-E", ciclo_numero=3)
    db.commit()

    assert _ciclos_consecutivos_deudor(db, "SAL-E") == 3

    db.query(Envio).filter(Envio.ciclo_id == ciclo.id).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id == ciclo.id).delete(synchronize_session=False)
    db.commit()


def test_confirmar_ciclo_marca_saldados_del_anterior(client, auth_headers, db, plantilla_default):
    """Integracion: confirmar un ciclo nuevo marca saldado a quien no reaparece."""
    import io
    import openpyxl
    from unittest.mock import patch, AsyncMock

    db.query(Ciclo).update({"activo": False})
    ciclo_ant = Ciclo(numero=9006, activo=True, creado_en=datetime.now(timezone.utc))
    db.add(ciclo_ant)
    db.flush()
    envio_que_salda = _make_envio(db, ciclo_ant, "SAL-F")
    envio_que_repite = _make_envio(db, ciclo_ant, "SAL-G")
    db.add(ClienteMaestro(clave_union="SAL-G", nombre="Repite", email="salg@mail.com",
                          actualizado_en=datetime.now(timezone.utc)))
    db.commit()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nro cliente", "nombre", "localidad", "monto"])
    ws.append(["SAL-G", "Repite", "CABA", 2000])
    buf = io.BytesIO()
    wb.save(buf)

    with patch("app.routers.ciclos.enviar_ciclo", new_callable=AsyncMock):
        r = client.post(
            "/ciclos/confirmar",
            files={"file": ("deudores.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
    assert r.status_code == 200

    db.expire_all()
    db.refresh(envio_que_salda)
    db.refresh(envio_que_repite)
    assert envio_que_salda.saldado_en is not None
    assert envio_que_repite.saldado_en is None

    ciclo_nuevo = db.query(Ciclo).filter(Ciclo.activo == True).first()
    db.query(Envio).filter(Envio.ciclo_id.in_([ciclo_ant.id, ciclo_nuevo.id])).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id.in_([ciclo_ant.id, ciclo_nuevo.id])).delete(synchronize_session=False)
    db.query(ClienteMaestro).filter(ClienteMaestro.clave_union == "SAL-G").delete(synchronize_session=False)
    db.commit()


def test_confirmar_asigna_racha_1_a_deudor_nuevo(client, auth_headers, db, plantilla_default):
    """Un deudor sin historial de Envios arranca con ciclo_numero == 1 (racha)."""
    import io
    import openpyxl
    from unittest.mock import patch, AsyncMock

    db.query(Ciclo).update({"activo": False})
    db.add(ClienteMaestro(clave_union="RN-A", nombre="Nuevo", email="rna@mail.com",
                          actualizado_en=datetime.now(timezone.utc)))
    db.commit()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nro cliente", "nombre", "localidad", "monto"])
    ws.append(["RN-A", "Nuevo", "CABA", 1000])
    buf = io.BytesIO()
    wb.save(buf)

    with patch("app.routers.ciclos.enviar_ciclo", new_callable=AsyncMock):
        r = client.post(
            "/ciclos/confirmar",
            files={"file": ("deudores.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
    assert r.status_code == 200

    ciclo_nuevo = db.query(Ciclo).filter(Ciclo.activo == True).first()
    envio_nuevo = (
        db.query(Envio)
        .filter(Envio.clave_union == "RN-A", Envio.ciclo_id == ciclo_nuevo.id)
        .first()
    )
    assert envio_nuevo is not None
    assert envio_nuevo.ciclo_numero == 1

    db.query(Envio).filter(Envio.ciclo_id == ciclo_nuevo.id).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id == ciclo_nuevo.id).delete(synchronize_session=False)
    db.query(ClienteMaestro).filter(ClienteMaestro.clave_union == "RN-A").delete(synchronize_session=False)
    db.commit()


def test_confirmar_incrementa_racha_del_que_repite(client, auth_headers, db, plantilla_default):
    """Un deudor con Envio previo sin saldar continua la racha (+1), no el numero global de ciclo."""
    import io
    import openpyxl
    from unittest.mock import patch, AsyncMock

    db.query(Ciclo).update({"activo": False})
    ciclo_ant = Ciclo(numero=9101, activo=True, creado_en=datetime.now(timezone.utc))
    db.add(ciclo_ant)
    db.flush()
    _make_envio(db, ciclo_ant, "RN-B", ciclo_numero=2)
    db.add(ClienteMaestro(clave_union="RN-B", nombre="Repite", email="rnb@mail.com",
                          actualizado_en=datetime.now(timezone.utc)))
    db.commit()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nro cliente", "nombre", "localidad", "monto"])
    ws.append(["RN-B", "Repite", "CABA", 2000])
    buf = io.BytesIO()
    wb.save(buf)

    with patch("app.routers.ciclos.enviar_ciclo", new_callable=AsyncMock):
        r = client.post(
            "/ciclos/confirmar",
            files={"file": ("deudores.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
    assert r.status_code == 200

    ciclo_nuevo = db.query(Ciclo).filter(Ciclo.activo == True).first()
    envio_nuevo = (
        db.query(Envio)
        .filter(Envio.clave_union == "RN-B", Envio.ciclo_id == ciclo_nuevo.id)
        .first()
    )
    assert envio_nuevo is not None
    assert envio_nuevo.ciclo_numero == 3

    db.query(Envio).filter(Envio.ciclo_id.in_([ciclo_ant.id, ciclo_nuevo.id])).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id.in_([ciclo_ant.id, ciclo_nuevo.id])).delete(synchronize_session=False)
    db.query(ClienteMaestro).filter(ClienteMaestro.clave_union == "RN-B").delete(synchronize_session=False)
    db.commit()


def test_confirmar_resetea_racha_tras_saldado(client, auth_headers, db, plantilla_default):
    """Un deudor cuyo ultimo Envio esta saldado arranca la racha de cero al reaparecer."""
    import io
    import openpyxl
    from unittest.mock import patch, AsyncMock

    db.query(Ciclo).update({"activo": False})
    ciclo_viejo = Ciclo(numero=9102, activo=False, creado_en=datetime.now(timezone.utc))
    db.add(ciclo_viejo)
    db.flush()
    _make_envio(db, ciclo_viejo, "RN-C", ciclo_numero=4, saldado_en=datetime.now(timezone.utc))
    db.add(ClienteMaestro(clave_union="RN-C", nombre="Saldado", email="rnc@mail.com",
                          actualizado_en=datetime.now(timezone.utc)))
    db.commit()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nro cliente", "nombre", "localidad", "monto"])
    ws.append(["RN-C", "Saldado", "CABA", 1500])
    buf = io.BytesIO()
    wb.save(buf)

    with patch("app.routers.ciclos.enviar_ciclo", new_callable=AsyncMock):
        r = client.post(
            "/ciclos/confirmar",
            files={"file": ("deudores.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
    assert r.status_code == 200

    ciclo_nuevo = db.query(Ciclo).filter(Ciclo.activo == True).first()
    envio_nuevo = (
        db.query(Envio)
        .filter(Envio.clave_union == "RN-C", Envio.ciclo_id == ciclo_nuevo.id)
        .first()
    )
    assert envio_nuevo is not None
    assert envio_nuevo.ciclo_numero == 1

    db.query(Envio).filter(Envio.ciclo_id.in_([ciclo_viejo.id, ciclo_nuevo.id])).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id.in_([ciclo_viejo.id, ciclo_nuevo.id])).delete(synchronize_session=False)
    db.query(ClienteMaestro).filter(ClienteMaestro.clave_union == "RN-C").delete(synchronize_session=False)
    db.commit()


def test_confirmar_asigna_racha_a_filtrados_y_sin_email(client, auth_headers, db, plantilla_default):
    """FILTRADO y SIN_EMAIL tambien son deudores: la racha se computa para ellos tambien."""
    import io
    import openpyxl
    from unittest.mock import patch, AsyncMock

    db.query(Ciclo).update({"activo": False})
    db.add(ClienteMaestro(clave_union="RN-D", nombre="Baja", email="rnd@mail.com",
                          prefiere_no_recibir_email=True,
                          actualizado_en=datetime.now(timezone.utc)))
    db.commit()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nro cliente", "nombre", "localidad", "monto"])
    ws.append(["RN-D", "Baja", "CABA", 1000])
    ws.append(["RN-E", "SinMaestro", "CABA", 1000])
    buf = io.BytesIO()
    wb.save(buf)

    with patch("app.routers.ciclos.enviar_ciclo", new_callable=AsyncMock):
        r = client.post(
            "/ciclos/confirmar",
            files={"file": ("deudores.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
    assert r.status_code == 200

    ciclo_nuevo = db.query(Ciclo).filter(Ciclo.activo == True).first()
    envio_filtrado = (
        db.query(Envio).filter(Envio.clave_union == "RN-D", Envio.ciclo_id == ciclo_nuevo.id).first()
    )
    envio_sin_email = (
        db.query(Envio).filter(Envio.clave_union == "RN-E", Envio.ciclo_id == ciclo_nuevo.id).first()
    )
    assert envio_filtrado is not None and envio_filtrado.ciclo_numero == 1
    assert envio_sin_email is not None and envio_sin_email.ciclo_numero == 1

    db.query(Envio).filter(Envio.ciclo_id == ciclo_nuevo.id).delete(synchronize_session=False)
    db.query(Ciclo).filter(Ciclo.id == ciclo_nuevo.id).delete(synchronize_session=False)
    db.query(ClienteMaestro).filter(ClienteMaestro.clave_union == "RN-D").delete(synchronize_session=False)
    db.commit()
