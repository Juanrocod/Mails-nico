# backend/tests/conftest.py
import os
from cryptography.fernet import Fernet

os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("SECRET_KEY", "test_secret_key_minimum_32_characters_here_ok")
os.environ.setdefault("ENCRYPTION_KEY", Fernet.generate_key().decode())
os.environ.setdefault("TOTP_ISSUER", "GestionMailsTest")
os.environ.setdefault("RATELIMIT_ENABLED", "false")

import pytest
import pyotp
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from fastapi.testclient import TestClient

from app.core.database import Base, get_db
from app.core.security import hash_password, generate_totp_secret
from app.models.user import User

_engine = create_engine(
    "sqlite:///:memory:",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)


@event.listens_for(_engine, "connect")
def _set_sqlite_pragma(dbapi_conn, _):
    cursor = dbapi_conn.cursor()
    cursor.execute("PRAGMA foreign_keys=ON")
    cursor.close()


_TestingSessionLocal = sessionmaker(_engine, autocommit=False, autoflush=False)


@pytest.fixture(scope="session", autouse=True)
def setup_test_database():
    Base.metadata.create_all(_engine)
    yield
    Base.metadata.drop_all(_engine)


@pytest.fixture
def db(setup_test_database):
    session = _TestingSessionLocal()
    try:
        yield session
        session.rollback()
    finally:
        session.close()


@pytest.fixture
def client(db):
    from app.main import app
    app.dependency_overrides[get_db] = lambda: db
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


@pytest.fixture
def test_user(db):
    totp_secret = generate_totp_secret()
    user = User(
        username=f"test_{os.urandom(4).hex()}",
        hashed_password=hash_password("SecurePass123!"),
        totp_secret=totp_secret,
        is_active=True,
    )
    db.add(user)
    db.flush()
    return user, totp_secret


@pytest.fixture
def auth_headers(client, test_user):
    user, totp_secret = test_user
    r = client.post(
        "/auth/login",
        json={"username": user.username, "password": "SecurePass123!"},
    )
    assert r.status_code == 200, f"Login failed: {r.text}"
    pending_token = r.json()["pending_token"]

    code = pyotp.TOTP(totp_secret).now()
    r = client.post(
        "/auth/verify-totp",
        json={"pending_token": pending_token, "code": code},
    )
    assert r.status_code == 200, f"TOTP verification failed: {r.text}"
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


@pytest.fixture
def make_valid_excel():
    import io
    import openpyxl
    from datetime import datetime
    from app.services.excel_parser import EXPECTED_COLUMNS

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(EXPECTED_COLUMNS.values())
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    row = {
        "cliente_nombre": "Test Cliente",
        "cliente_email": "cliente@test.com",
        "cuenta_comitente": "12345",
        "cuenta_cotapartista": "67890",
        "instrumento": "AL30",
        "tipo": "COMPRA",
        "cantidad": 100.0,
        "precio": 70.50,
        "moneda": "USD",
        "liquidacion": "24HS",
        "fecha_operacion": datetime(2026, 6, 14, 10, 30),
    }
    for col_idx, key in enumerate(EXPECTED_COLUMNS.keys(), 1):
        ws.cell(row=2, column=col_idx, value=row[key])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def seeded_borrador_minuta(client, auth_headers, make_valid_excel):
    """Upload a valid Excel and return the UUID string of the first BORRADOR minuta."""
    r = client.post(
        "/uploads/excel",
        files={"file": ("ops.xlsx", make_valid_excel,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert r.status_code == 201, f"Upload failed: {r.text}"
    minutas = r.json()["minutas"]
    assert len(minutas) >= 1
    return minutas[0]["id"]
