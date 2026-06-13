import pytest
from uuid import uuid4
from datetime import datetime

from app.models.order import Orden, ExcelUpload, EstadoMinuta, TipoOperacion, CondicionLiquidacion
from app.models.user import User
from app.models.audit import AuditEvent, AccionAudit
from app.core.security import hash_password
from app.services.audit import record_event, get_events_for_orden, export_audit_trail_excel


def make_user(db):
    user = User(
        username=f"u_{uuid4().hex[:8]}",
        hashed_password=hash_password("pass"),
        totp_secret="JBSWY3DPEHPK3PXP",
    )
    db.add(user)
    db.flush()
    return user


def make_orden(db):
    user = make_user(db)
    upload = ExcelUpload(
        usuario_id=user.id,
        nombre_archivo="test.xlsx",
        total_ordenes=1,
        ordenes_validas=1,
        ordenes_con_error=0,
    )
    db.add(upload)
    db.flush()
    orden = Orden(
        excel_upload_id=upload.id,
        cliente_nombre="Test",
        cliente_email="t@t.com",
        cuenta_comitente="111",
        cuenta_cotapartista="222",
        instrumento="AL30",
        tipo=TipoOperacion.COMPRA,
        cantidad=100,
        precio=70.0,
        moneda="USD",
        liquidacion=CondicionLiquidacion.HS24,
        fecha_operacion=datetime(2026, 6, 13, 10, 0),
        estado=EstadoMinuta.BORRADOR,
        texto_minuta="Texto",
        texto_editado=False,
    )
    db.add(orden)
    db.flush()
    return orden, user


def test_record_event_creates_audit_event(db):
    orden, user = make_orden(db)
    event = record_event(orden.id, user.id, AccionAudit.CREADA, "127.0.0.1", db)
    db.flush()
    assert event.id is not None
    assert event.orden_id == orden.id
    assert event.accion == AccionAudit.CREADA
    assert event.ip_origen == "127.0.0.1"


def test_record_event_with_null_usuario(db):
    orden, _ = make_orden(db)
    event = record_event(orden.id, None, AccionAudit.ALERTA_GENERADA, None, db)
    db.flush()
    assert event.usuario_id is None
    assert event.ip_origen is None


def test_record_event_with_detalle(db):
    orden, user = make_orden(db)
    detalle = {"campos_editados": ["texto_minuta"], "texto_anterior": "viejo"}
    event = record_event(orden.id, user.id, AccionAudit.EDITADA, "127.0.0.1", db, detalle=detalle)
    db.flush()
    assert event.detalle == detalle


def test_get_events_for_orden_returns_in_order(db):
    orden, user = make_orden(db)
    record_event(orden.id, user.id, AccionAudit.CREADA, "1.1.1.1", db)
    record_event(orden.id, user.id, AccionAudit.EDITADA, "1.1.1.1", db)
    record_event(orden.id, user.id, AccionAudit.APROBADA, "1.1.1.1", db)
    db.flush()
    events = get_events_for_orden(orden.id, db)
    assert len(events) == 3
    acciones = [e.accion for e in events]
    assert acciones == [AccionAudit.CREADA, AccionAudit.EDITADA, AccionAudit.APROBADA]


def test_get_events_for_orden_empty(db):
    orden, _ = make_orden(db)
    events = get_events_for_orden(orden.id, db)
    assert events == []


def test_export_audit_trail_excel_returns_bytes(db):
    orden, user = make_orden(db)
    record_event(orden.id, user.id, AccionAudit.CREADA, "127.0.0.1", db)
    db.flush()
    content = export_audit_trail_excel(orden.id, db)
    assert isinstance(content, bytes)
    assert len(content) > 100  # non-empty xlsx


def test_export_audit_trail_excel_is_valid_xlsx(db):
    import io
    import openpyxl
    orden, user = make_orden(db)
    record_event(orden.id, user.id, AccionAudit.CREADA, "127.0.0.1", db)
    db.flush()
    content = export_audit_trail_excel(orden.id, db)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    # Header row
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert "Timestamp" in headers
    assert "Acción" in headers
    # At least one data row
    assert ws.max_row >= 2
