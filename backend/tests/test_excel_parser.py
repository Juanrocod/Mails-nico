import io
import openpyxl
import pytest
from app.services.excel_parser import parse_deudores, parse_maestro, ExcelParseError


def _make_excel(headers: list, rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_deudores_basico():
    data = _make_excel(
        ["nro cliente", "nombre", "localidad", "monto"],
        [["C001", "Consorcio Test", "CABA", 5000.00]],
    )
    rows = parse_deudores(data)
    assert len(rows) == 1
    assert rows[0].clave_union == "C001"
    assert rows[0].nombre == "Consorcio Test"
    assert rows[0].localidad == "CABA"
    assert float(rows[0].monto) == 5000.00


def test_parse_deudores_alias_detalle():
    """Acepta 'detalle del nombre' como alias de nombre."""
    data = _make_excel(
        ["nro cliente", "detalle del nombre", "localidad", "monto"],
        [["C002", "Otro Consorcio", "GBA", 3000.50]],
    )
    rows = parse_deudores(data)
    assert rows[0].nombre == "Otro Consorcio"


def test_parse_deudores_omite_monto_cero():
    data = _make_excel(
        ["nro cliente", "nombre", "localidad", "monto"],
        [["C003", "Consorcio Cero", "CABA", 0]],
    )
    rows = parse_deudores(data)
    assert len(rows) == 0


def test_parse_deudores_columna_faltante_lanza_error():
    data = _make_excel(["nombre", "monto"], [["Consorcio", 100]])
    with pytest.raises(ExcelParseError, match="clave_union"):
        parse_deudores(data)


def test_parse_maestro_basico():
    data = _make_excel(
        ["nro cliente", "nombre", "email", "localidad"],
        [["C001", "Consorcio Test", "test@mail.com", "CABA"]],
    )
    rows = parse_maestro(data)
    assert len(rows) == 1
    assert rows[0].email == "test@mail.com"


def test_parse_maestro_email_vacio_es_none():
    data = _make_excel(
        ["nro cliente", "nombre", "email", "localidad"],
        [["C001", "Consorcio Test", "", "CABA"]],
    )
    rows = parse_maestro(data)
    assert rows[0].email is None
